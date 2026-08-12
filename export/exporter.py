from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import os

from scraper.source_tracking import enrich_source_tracking

HEADERS = [
    'Date Applied', 'Company', 'Job Title', 'Job link', 'Status', 'Location', 'Work Type',
    'Salary Range', 'Follow-up', 'Found On', 'Application Portal'
]


def export_jobs_to_xlsx(jobs, outdir='exports'):
    os.makedirs(outdir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Jobs'
    ws.append(HEADERS)
    for j in jobs:
        tracked = enrich_source_tracking(j, j.get('job_link'))
        row = [
            tracked.get('date_applied', ''),
            tracked.get('company', ''),
            tracked.get('job_title', ''),
            tracked.get('job_link', ''),
            tracked.get('status', ''),
            tracked.get('location', ''),
            tracked.get('work_type', ''),
            tracked.get('salary', ''),
            tracked.get('follow_up', ''),
            tracked.get('found_on', 'N/A'),
            tracked.get('application_portal', 'Company Website'),
        ]
        ws.append(row)
    # auto-width (simple)
    for i, col in enumerate(ws.columns, 1):
        maxlen = 0
        for cell in col:
            try:
                val = str(cell.value or '')
            except:
                val = ''
            if len(val) > maxlen:
                maxlen = len(val)
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen, 10), 60)

    fname = f"job_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(outdir, fname)
    wb.save(out_path)
    return out_path
