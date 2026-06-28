from __future__ import annotations

from copy import copy
from datetime import date, datetime
import os
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table


CANONICAL_HEADERS = [
    "Date Applied",
    "Company",
    "Job Title",
    "Job link",
    "Status",
    "Location",
    "Work Type",
    "Salary Range",
    "Follow-up",
    "Source",
]

HEADER_ALIASES = {
    "Date Applied": {"dateapplied", "applieddate", "applicationdate", "date"},
    "Company": {"company", "companyname", "employer", "organization"},
    "Job Title": {"jobtitle", "title", "position", "role"},
    "Job link": {"joblink", "joburl", "url", "link", "postinglink", "applicationlink"},
    "Status": {"status", "applicationstatus"},
    "Location": {"location", "joblocation"},
    "Work Type": {"worktype", "workstyle", "remotehybridonsite", "remote"},
    "Salary Range": {"salaryrange", "salary", "pay", "compensation"},
    "Follow-up": {"followup", "followupdate", "followupdate"},
    "Source": {"source", "jobboard", "platform", "website"},
}

JOB_KEYS = {
    "Date Applied": "date_applied",
    "Company": "company",
    "Job Title": "job_title",
    "Job link": "job_link",
    "Status": "status",
    "Location": "location",
    "Work Type": "work_type",
    "Salary Range": "salary",
    "Follow-up": "follow_up",
    "Source": "source",
}


def append_jobs_to_workbook(file_obj, filename: str, jobs: list[dict], outdir: str = "exports") -> tuple[str, dict]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Upload an .xlsx or .xlsm Excel workbook.")

    keep_vba = suffix == ".xlsm"
    workbook = load_workbook(file_obj, keep_vba=keep_vba)
    worksheet, header_row = _choose_worksheet(workbook)
    columns = _ensure_headers(worksheet, header_row)
    existing_links = _existing_links(worksheet, columns["Job link"], header_row)

    added = 0
    skipped = 0
    last_row = header_row
    for job in jobs:
        link = _clean_text(job.get("job_link"))
        if link and link in existing_links:
            skipped += 1
            continue
        row_number = _first_empty_row(worksheet, columns["Job link"], header_row)
        if row_number > header_row + 1:
            _copy_row_style(worksheet, row_number - 1, row_number)
        _write_job_row(worksheet, row_number, columns, job)
        if link:
            existing_links.add(link)
        added += 1
        last_row = max(last_row, row_number)

    _extend_tables(worksheet, last_row)
    _tidy_columns(worksheet, columns)

    os.makedirs(outdir, exist_ok=True)
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(filename).stem or "job_tracker").strip("._")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(outdir, f"{stem}_with_jobs_{timestamp}{suffix}")
    workbook.save(out_path)
    return out_path, {"added": added, "skipped": skipped, "sheet": worksheet.title}


def _normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _clean_text(value: object) -> str:
    return str(value or "").strip()


def _canonical_header(value: object) -> str | None:
    normalized = _normalize(value)
    for header, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return header
    return None


def _header_map(worksheet, row_number: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for cell in worksheet[row_number]:
        header = _canonical_header(cell.value)
        if header and header not in columns:
            columns[header] = cell.column
    return columns


def _choose_worksheet(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().casefold() == "applications":
            worksheet = workbook[sheet_name]
            return worksheet, _best_header_row(worksheet)

    best = (None, 1, -1)
    for worksheet in workbook.worksheets:
        row_number, score = _score_header_rows(worksheet)
        if score > best[2]:
            best = (worksheet, row_number, score)
    worksheet = best[0] or workbook.active
    return worksheet, best[1] if best[2] > 0 else 1


def _best_header_row(worksheet) -> int:
    return _score_header_rows(worksheet)[0]


def _score_header_rows(worksheet) -> tuple[int, int]:
    best_row = 1
    best_score = -1
    max_row = min(max(worksheet.max_row, 1), 10)
    for row_number in range(1, max_row + 1):
        score = len(_header_map(worksheet, row_number))
        if score > best_score:
            best_row = row_number
            best_score = score
    return best_row, best_score


def _ensure_headers(worksheet, header_row: int) -> dict[str, int]:
    columns = _header_map(worksheet, header_row)
    next_column = max(worksheet.max_column, 1) + 1
    if not any(_clean_text(cell.value) for cell in worksheet[header_row]):
        next_column = 1

    for header in CANONICAL_HEADERS:
        if header in columns:
            continue
        column = next_column
        next_column += 1
        cell = worksheet.cell(row=header_row, column=column, value=header)
        _copy_cell_style(worksheet.cell(row=header_row, column=max(1, column - 1)), cell)
        columns[header] = column
    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1).coordinate
    return columns


def _existing_links(worksheet, link_column: int, header_row: int) -> set[str]:
    links = set()
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=link_column)
        link = _clean_text(cell.hyperlink.target if cell.hyperlink else cell.value)
        if link and link.casefold() != "open job":
            links.add(link)
        elif cell.hyperlink and cell.hyperlink.target:
            links.add(_clean_text(cell.hyperlink.target))
    return links


def _first_empty_row(worksheet, link_column: int, header_row: int) -> int:
    for row_number in range(header_row + 1, worksheet.max_row + 2):
        cell = worksheet.cell(row=row_number, column=link_column)
        link = _clean_text(cell.hyperlink.target if cell.hyperlink else cell.value)
        if not link:
            return row_number
    return worksheet.max_row + 1


def _write_job_row(worksheet, row_number: int, columns: dict[str, int], job: dict) -> None:
    for header, key in JOB_KEYS.items():
        column = columns.get(header)
        if not column:
            continue
        value = job.get(key, "")
        if header == "Date Applied":
            value = _excel_date(value)
        elif header == "Follow-up" and not value:
            value = ""
        elif header in {"Company", "Job Title", "Location", "Work Type", "Salary Range"} and not value:
            value = "n/a"
        elif header == "Source" and not value:
            value = "Company Website"
        worksheet.cell(row=row_number, column=column, value=value)

    link = _clean_text(job.get("job_link"))
    if link and columns.get("Job link"):
        link_cell = worksheet.cell(row=row_number, column=columns["Job link"])
        link_cell.value = link
        link_cell.hyperlink = link
        link_cell.style = "Hyperlink"

    for header in ("Date Applied", "Follow-up"):
        column = columns.get(header)
        if column:
            worksheet.cell(row=row_number, column=column).number_format = "mm/dd/yyyy"


def _excel_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    for pattern in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return date.today()


def _copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    if source_row < 1 or source_row == target_row:
        return
    for column in range(1, worksheet.max_column + 1):
        _copy_cell_style(worksheet.cell(row=source_row, column=column), worksheet.cell(row=target_row, column=column))
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def _extend_tables(worksheet, last_row: int) -> None:
    for table in worksheet.tables.values():
        if not isinstance(table, Table):
            continue
        start, end = table.ref.split(":")
        end_column = re.match(r"[A-Z]+", end).group(0)
        end_row = int(re.search(r"\d+", end).group(0))
        if last_row > end_row:
            table.ref = f"{start}:{end_column}{last_row}"


def _tidy_columns(worksheet, columns: dict[str, int]) -> None:
    preferred = {
        "Job link": 60,
        "Salary Range": 24,
        "Job Title": 34,
        "Company": 24,
        "Location": 28,
        "Work Type": 14,
        "Source": 18,
    }
    for header, width in preferred.items():
        column = columns.get(header)
        if not column:
            continue
        letter = get_column_letter(column)
        current = worksheet.column_dimensions[letter].width or 0
        worksheet.column_dimensions[letter].width = max(current, width)
