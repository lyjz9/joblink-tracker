#!/usr/bin/env python
"""Process pending job links from an Excel workbook."""

from __future__ import annotations

import argparse
from copy import copy
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table

from scraper.source_tracking import enrich_source_tracking

INPUT_HEADERS = [
    "Job Link",
    "Notes",
    "Process Status",
    "Processed At",
    "Error Message",
]

APPLICATION_HEADERS = [
    "Date Applied",
    "Company",
    "Job Title",
    "Job link",
    "Status",
    "Location",
    "Work Type",
    "Salary Range",
    "Follow-up",
    "Application Portal",
]

SHEET_WIDTHS = {
    "Input": [70, 35, 24, 22, 55],
    "Applications": [15, 26, 38, 70, 18, 28, 16, 24, 15, 22],
}


def normalize_header(value: object) -> str:
    return str(value or "").strip().casefold()


def header_map(worksheet) -> dict[str, int]:
    return {
        normalize_header(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value
    }


def ensure_headers(worksheet, headers: list[str]) -> None:
    existing = header_map(worksheet)
    legacy_source_column = None
    if worksheet.title == "Applications" and "source" in existing:
        if "application portal" not in existing:
            legacy_source_column = existing["source"]
            worksheet.cell(row=1, column=existing["source"], value="Application Portal")
        existing = header_map(worksheet)
    next_column = max(existing.values(), default=0) + 1
    for header in headers:
        if normalize_header(header) not in existing:
            worksheet.cell(row=1, column=next_column, value=header)
            existing[normalize_header(header)] = next_column
            next_column += 1
    worksheet.freeze_panes = "A2"
    if worksheet.tables:
        worksheet.auto_filter.ref = None
    else:
        worksheet.auto_filter.ref = (
            f"A1:{worksheet.cell(1, max(existing.values(), default=len(headers))).coordinate}"
        )
    worksheet.row_dimensions[1].height = 24
    widths = SHEET_WIDTHS.get(worksheet.title, [])
    for header, width in zip(headers, widths):
        column = existing[normalize_header(header)]
        cell = worksheet.cell(row=1, column=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[cell.column_letter].width = width
    if worksheet.title == "Applications":
        migrate_application_portals(
            worksheet,
            legacy_source_column=legacy_source_column,
        )


def migrate_application_portals(worksheet, *, legacy_source_column: int | None = None) -> None:
    columns = header_map(worksheet)
    link_column = columns.get("job link")
    portal_column = columns.get("application portal")
    if not link_column or not portal_column:
        return
    for row in range(2, worksheet.max_row + 1):
        link_cell = worksheet.cell(row=row, column=link_column)
        link = clean_url(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value)
        company = worksheet.cell(row=row, column=columns.get("company", 1)).value
        title = worksheet.cell(row=row, column=columns.get("job title", 1)).value
        if not link and not company and not title:
            continue
        portal = worksheet.cell(row=row, column=portal_column).value
        if legacy_source_column:
            tracked = enrich_source_tracking({
                "job_link": link,
                "source": worksheet.cell(row=row, column=legacy_source_column).value,
            }, link)
        else:
            tracked = enrich_source_tracking({
                "job_link": link,
                "application_portal": portal,
                "source": portal,
            }, link)
        if legacy_source_column or not portal:
            worksheet.cell(row=row, column=portal_column, value=tracked["application_portal"])


def get_or_create_sheet(workbook, name: str, headers: list[str]):
    if name in workbook.sheetnames:
        worksheet = workbook[name]
    elif name == "Applications":
        worksheet = next(
            (
                sheet
                for sheet in workbook.worksheets
                if "job link" in header_map(sheet)
            ),
            None,
        )
        if worksheet is None:
            worksheet = workbook.create_sheet(name)
        else:
            worksheet.title = name
    else:
        worksheet = workbook.create_sheet(name)
    ensure_headers(worksheet, headers)
    return worksheet


def clean_url(value: object) -> str:
    return str(value or "").strip()


def existing_job_links(worksheet) -> set[str]:
    columns = header_map(worksheet)
    link_column = columns.get("job link")
    if not link_column:
        return set()
    links = set()
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=link_column)
        link = clean_url(cell.hyperlink.target if cell.hyperlink else cell.value)
        if link:
            links.add(link)
    return links


def excel_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return date.today()


def result_status(result: dict) -> str:
    if result.get("error"):
        return "Error"
    required = ("company", "job_title", "location")
    if any(str(result.get(key, "")).strip().casefold() in {"", "n/a"} for key in required):
        return "Needs Manual Review"
    return "Done"


def append_application(worksheet, result: dict, original_url: str) -> None:
    result = enrich_source_tracking(result, result.get("job_link") or original_url)
    applied = excel_date(result.get("date_applied"))
    follow_up = result.get("follow_up", "")
    values = {
        "date applied": applied,
        "company": result.get("company", "n/a"),
        "job title": result.get("job_title", "n/a"),
        "job link": result.get("job_link") or original_url,
        "status": result.get("status", ""),
        "location": result.get("location", "n/a"),
        "work type": result.get("work_type", "n/a"),
        "salary range": result.get("salary", "n/a"),
        "follow-up": follow_up,
        "application portal": result.get("application_portal", "Company Website"),
    }
    columns = header_map(worksheet)
    link_column = columns.get("job link")
    new_row = next(
        (
            row
            for row in range(2, worksheet.max_row + 1)
            if not clean_url(worksheet.cell(row=row, column=link_column).value)
        ),
        worksheet.max_row + 1,
    )
    style_source = next(
        (
            row
            for row in range(new_row - 1, 1, -1)
            if any(
                clean_url(worksheet.cell(row=row, column=column).value)
                for column in columns.values()
            )
        ),
        None,
    )
    if style_source is None:
        style_source = next(
            (
                row
                for row in range(new_row - 1, 1, -1)
                if any(
                    worksheet.cell(row=row, column=column).has_style
                    for column in range(1, worksheet.max_column + 1)
                )
            ),
            None,
        )
    if style_source:
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(row=style_source, column=column)
            target = worksheet.cell(row=new_row, column=column)
            if source.has_style:
                target._style = copy(source._style)
        worksheet.row_dimensions[new_row].height = worksheet.row_dimensions[style_source].height
    for header, value in values.items():
        column = columns.get(header)
        if column:
            worksheet.cell(row=new_row, column=column, value=value)
    if link_column:
        link_cell = worksheet.cell(row=new_row, column=link_column)
        link_cell.value = "Open job"
        link_cell.hyperlink = original_url
        link_cell.style = "Hyperlink"
    for header in ("date applied", "follow-up"):
        column = columns.get(header)
        if column:
            worksheet.cell(row=new_row, column=column).number_format = "mm/dd/yyyy"
    _extend_application_tables(worksheet, new_row, columns)


def _extend_application_tables(worksheet, new_row: int, columns: dict[str, int]) -> None:
    link_column = columns.get("job link", 0)
    for table in worksheet.tables.values():
        if not isinstance(table, Table):
            continue
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        if min_row != 1 or not (min_column <= link_column <= max_column):
            continue
        table.ref = (
            f"{get_column_letter(min_column)}{min_row}:"
            f"{get_column_letter(max_column)}{max(max_row, new_row)}"
        )


def process_workbook(path: Path) -> tuple[int, int, int]:
    keep_vba = path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(path, keep_vba=keep_vba)
    applications = get_or_create_sheet(workbook, "Applications", APPLICATION_HEADERS)
    input_sheet = get_or_create_sheet(workbook, "Input", INPUT_HEADERS)

    input_columns = header_map(input_sheet)
    known_links = existing_job_links(applications)
    processed = duplicates = errors = 0
    scraper = None

    for row in range(2, input_sheet.max_row + 1):
        link = clean_url(input_sheet.cell(row, input_columns["job link"]).value)
        status = str(input_sheet.cell(row, input_columns["process status"]).value or "").strip()
        if not link or status.casefold() not in {"", "pending"}:
            continue

        if link in known_links:
            input_sheet.cell(row, input_columns["process status"], "Duplicate")
            input_sheet.cell(row, input_columns["processed at"], datetime.now())
            duplicates += 1
            continue

        try:
            if scraper is None:
                from scraper.browser_scraper_v2 import parse_job_with_browser

                scraper = parse_job_with_browser
            result = scraper(link)
            result = enrich_source_tracking(result, link)
            status = result_status(result)
            if status != "Error":
                append_application(applications, result, link)
                known_links.add(link)
                processed += 1
            else:
                errors += 1
            input_sheet.cell(row, input_columns["process status"], status)
            input_sheet.cell(row, input_columns["processed at"], datetime.now())
            input_sheet.cell(row, input_columns["error message"], result.get("error", ""))
        except Exception as exc:
            input_sheet.cell(row, input_columns["process status"], "Error")
            input_sheet.cell(row, input_columns["processed at"], datetime.now())
            input_sheet.cell(row, input_columns["error message"], str(exc))
            errors += 1

        workbook.save(path)

    workbook.save(path)
    return processed, duplicates, errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Process pending job links from Excel.")
    parser.add_argument("workbook", help="Path to the .xlsx or .xlsm tracker workbook")
    args = parser.parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        print(f"I could not find that workbook: {workbook_path}")
        return 1

    processed, duplicates, errors = process_workbook(workbook_path)
    print(f"Done. Added: {processed}. Duplicates: {duplicates}. Errors: {errors}.")
    return 0 if errors == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
