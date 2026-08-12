from openpyxl import load_workbook

from export.exporter import HEADERS, export_jobs_to_xlsx


def test_new_tracker_separates_discovery_and_application_portal(tmp_path):
    output_path = export_jobs_to_xlsx([{
        "company": "Example Company",
        "job_title": "Analyst",
        "job_link": "https://jobs.ashbyhq.com/example/123",
        "source": "LinkedIn",
    }], outdir=tmp_path)

    worksheet = load_workbook(output_path).active
    assert [cell.value for cell in worksheet[1]] == HEADERS
    assert worksheet.cell(2, HEADERS.index("Found On") + 1).value == "LinkedIn"
    assert worksheet.cell(2, HEADERS.index("Application Portal") + 1).value == "Ashby"
