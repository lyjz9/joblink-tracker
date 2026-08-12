from __future__ import annotations

import io
import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

import scraper.app as app_module


@pytest.fixture()
def client(tmp_path):
    app = app_module.create_app("testing", {
        "LOG_DIR": tmp_path,
        "RATE_LIMIT_UPLOAD": 10,
        "RATE_WINDOW_SECONDS": 600,
        "MAX_EXPORT_JOBS": 100,
        "MAX_WORKBOOK_UNCOMPRESSED_BYTES": 80 * 1024 * 1024,
        "MAX_WORKBOOK_ARCHIVE_MEMBERS": 5000,
    })
    try:
        with app.test_client() as test_client:
            yield test_client
    finally:
        app_module.shutdown_app(app)


def test_append_workbook_rejects_fake_excel_file(client):
    response = client.post(
        "/append-workbook",
        data={
            "workbook": (io.BytesIO(b"not a workbook"), "tracker.xlsx"),
            "jobs": json.dumps([{"company": "Example"}]),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "readable Excel workbook" in response.get_json()["error"]


def test_append_workbook_returns_valid_xlsx(client):
    template = Path(__file__).resolve().parents[1] / "templates" / "linc_tracker_template.xlsx"
    jobs = [{
        "date_applied": "07/12/2026",
        "company": "Example Company",
        "job_title": "Data Analyst",
        "job_link": "https://example.wd5.myworkdayjobs.com/job/route-test?source=LinkedIn",
        "location": "New York, NY",
        "work_type": "Hybrid",
        "salary": "n/a",
        "found_on": "LinkedIn",
        "application_portal": "Workday",
        "source": "Workday",
    }]
    response = client.post(
        "/append-workbook",
        data={
            "workbook": (io.BytesIO(template.read_bytes()), template.name),
            "jobs": json.dumps(jobs),
            "duplicate_mode": "skip",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    assert response.headers["X-JobLink-Added"] == "1"
    with ZipFile(io.BytesIO(response.data)) as workbook:
        assert "xl/workbook.xml" in workbook.namelist()
    output = load_workbook(io.BytesIO(response.data))
    worksheet = output["Applications"]
    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value
    }
    assert "Source" not in headers
    assert headers.keys() >= {"Found On", "Application Portal"}
    row = next(
        row_number
        for row_number in range(2, worksheet.max_row + 1)
        if worksheet.cell(row_number, headers["Company"]).value == "Example Company"
    )
    assert worksheet.cell(row, headers["Found On"]).value == "LinkedIn"
    assert worksheet.cell(row, headers["Application Portal"]).value == "Workday"
