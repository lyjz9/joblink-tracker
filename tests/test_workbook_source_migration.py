from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from export.workbook_appender import append_jobs_to_workbook


def test_legacy_source_column_is_renamed_to_portal_without_growing_the_table(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Applications"
    worksheet.append([
        "Date Applied", "Company", "Job Title", "Job link", "Status",
        "Location", "Work Type", "Salary Range", "Follow-up", "Source",
    ])
    worksheet.append([
        "08/01/2026", "Example Company", "Analyst",
        "https://example.wd5.myworkdayjobs.com/job/Analyst_R123",
        "Applied", "New York, NY", "Hybrid", "n/a", "", "LinkedIn",
    ])
    table = Table(displayName="ApplicationsTable", ref="A1:J2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    output_path, _summary = append_jobs_to_workbook(
        source,
        "legacy.xlsx",
        [],
        outdir=tmp_path,
    )
    migrated = load_workbook(output_path)
    worksheet = migrated["Applications"]
    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value
    }

    assert "Source" not in headers
    assert worksheet.cell(2, headers["Application Portal"]).value == "Workday"
    assert worksheet.tables["ApplicationsTable"].ref == "A1:J2"


def test_append_preserves_data_row_formatting_and_legacy_found_on_column(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Applications"
    worksheet.append([
        "Date Applied", "Company", "Job Title", "Job link", "Status",
        "Location", "Work Type", "Salary Range", "Follow-up",
        "Application Portal", "Found On", "Notes",
    ])
    worksheet.append([
        "08/01/2026", "Existing Company", "Analyst", "Open job", "Applied",
        "New York, NY", "Hybrid", "$80,000", "", "Workday", "LinkedIn", "Keep this formula",
    ])
    worksheet.cell(2, 4).hyperlink = "https://example.wd5.myworkdayjobs.com/job/Existing_R123"
    worksheet.cell(2, 2).font = Font(name="Calibri", bold=True, color="FFFFFF")
    worksheet.cell(2, 2).fill = PatternFill("solid", fgColor="1F4E78")
    worksheet.cell(2, 2).alignment = Alignment(horizontal="center")
    worksheet.cell(2, 2).border = Border(bottom=Side(style="thin", color="FFFFFF"))
    worksheet.cell(2, 2).number_format = "@"
    worksheet.row_dimensions[2].height = 28
    worksheet.cell(2, 12, "=CONCAT(A2,\"-\",B2)")
    table = Table(displayName="ApplicationsTable", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    output_path, summary = append_jobs_to_workbook(
        source,
        "formatted.xlsx",
        [
            {
                "company": "New Company",
                "job_title": "Coordinator",
                "job_link": "https://jobs.lever.co/example/123",
                "location": "Remote",
                "work_type": "Remote",
                "salary": "$90,000",
            },
            {
                "company": "Next Company",
                "job_title": "Specialist",
                "job_link": "https://jobs.ashbyhq.com/example/456",
                "location": "Boston, MA",
                "work_type": "Hybrid",
                "salary": "$95,000",
            },
        ],
        outdir=tmp_path,
    )

    assert summary["added"] == 2
    reopened = load_workbook(output_path)
    sheet = reopened["Applications"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert "Found On" in headers
    assert sheet.cell(3, headers["Found On"]).value is None
    assert sheet.cell(4, headers["Found On"]).value is None
    assert sheet.cell(3, headers["Application Portal"]).value == "Lever"
    assert sheet.cell(4, headers["Application Portal"]).value == "Ashby"
    assert sheet.cell(2, 12).value == '=CONCAT(A2,"-",B2)'
    assert sheet.row_dimensions[3].height == 28
    for row in (3, 4):
        cell = sheet.cell(row, headers["Company"])
        source_cell = sheet.cell(2, headers["Company"])
        assert cell.font.bold == source_cell.font.bold
        assert cell.font.color.rgb == source_cell.font.color.rgb
        assert cell.fill.fill_type == source_cell.fill.fill_type
        assert cell.fill.fgColor.rgb == source_cell.fill.fgColor.rgb
        assert cell.alignment.horizontal == source_cell.alignment.horizontal
        assert cell.border.bottom.style == source_cell.border.bottom.style
        assert cell.number_format == source_cell.number_format
    assert sheet.tables["ApplicationsTable"].ref == "A1:K4"
