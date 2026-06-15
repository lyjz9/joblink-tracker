from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import os

HEADERS = [
    'Date Applied', 'Company', 'Job Title', 'Job link', 'Status', 'Location', 'Work Type',
    'Salary Range', 'Follow-up', 'Source'
]


def export_jobs_to_xlsx(jobs, outdir='exports'):
    os.makedirs(outdir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Jobs'
    ws.append(HEADERS)
    for j in jobs:
        row = [
            j.get('date_applied', ''),
            j.get('company', ''),
            j.get('job_title', ''),
            j.get('job_link', ''),
            j.get('status', ''),
            j.get('location', ''),
            j.get('work_type', ''),
            j.get('salary', ''),
            j.get('follow_up', ''),
            j.get('source', '')
        ]
        ws.append(row)
    # auto-width (simple)
    for i, col in enumerate(ws.columns, 1):
        maxlen = 0
        for cell in col:
            try:
                val = str(cell.value or '')
            except:
                val = ''
            if len(val) > maxlen:
                maxlen = len(val)
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen, 10), 60)

    fname = f"job_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(outdir, fname)
    wb.save(out_path)
    return out_path
